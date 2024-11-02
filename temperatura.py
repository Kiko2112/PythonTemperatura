import tkinter as tk
from tkinter import messagebox
import requests
from bs4 import BeautifulSoup
from datetime import datetime
import openpyxl
import os

class WeatherApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Captador de Temperatura - São Paulo")
        
        self.root.geometry("400x200")
        self.root.configure(bg='#f0f0f0')
        
        self.button = tk.Button(
            root,
            text="Buscar previsão",
            command=self.coletar_dados,
            bg='#4CAF50',
            fg='white',
            font=('Arial', 12),
            padx=20,
            pady=10
        )
        self.button.pack(expand=True)
        
        self.info_label = tk.Label(
            root,
            text="",
            bg='#f0f0f0',
            font=('Arial', 10)
        )
        self.info_label.pack(pady=10)

    def coletar_dados(self):
        try:
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
            }
            
            url = "https://www.climatempo.com.br/previsao-do-tempo/agora/cidade/558/saopaulo-sp"
            response = requests.get(url, headers=headers)
            soup = BeautifulSoup(response.text, 'html.parser')

            temperatura = (soup.find('span', class_='-bold -gray-dark-2 -font-55 _margin-l-20 _center').text.strip()
            if soup.find('span', class_='-bold -gray-dark-2 -font-55 _margin-l-20 _center') else 'N/A')
            
            umidade = (soup.find('span', class_='-gray-light').text.strip()
            if soup.find('span', class_='-gray-light') else 'N/A')
            
            self.info_label.config(
                text=f"Temperatura: {temperatura}\nUmidade: {umidade}"
            )
            
            self.save_to_excel(temperatura, umidade)
            
            messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao capturar os dados: {str(e)}\n")

    def save_to_excel(self, temperatura, umidade):
        try:
            filename = "temperatura_sp.xlsx"
            
            if not os.path.exists(filename):
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.append(["Data/Hora", "Temperatura", "Umidade"])
            else:
                wb = openpyxl.load_workbook(filename)
                ws = wb.active
            
            current_time = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            ws.append([current_time, temperatura, umidade])
            
            for column in ws.columns:
                max_length = 0
                column_letter = openpyxl.utils.get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filename)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar no Excel: {str(e)}")

if __name__ == "__main__":
    root = tk.Tk()
    app = WeatherApp(root)
    root.mainloop()